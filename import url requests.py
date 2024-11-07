import requests
from bs4 import BeautifulSoup
import nltk
from nltk.tokenize import sent_tokenize
from openpyxl import Workbook, load_workbook
import datetime
import os

# 指定文件路径
file_path = r"E:\\IME-S1\\SD5913\\asm4-taro\\data_collect"
file_name = "links of keyword_20241107_003107-emitons-p5.xlsx"

# 加载Excel文件
wb = load_workbook(os.path.join(file_path, file_name))
ws = wb.active

# 获取所有URL并去除重复
urls = list(set([cell.value for cell in ws['A'] if cell.value]))

# 创建文件保存路径
save_path = os.path.join(os.getcwd(), "data_outcome")
if not os.path.exists(save_path):
    os.makedirs(save_path)

# 遍历每个URL
for url in urls:
    try:
        # 检查URL是否以http或https开头
        if not url.startswith('http'):
            print(f"跳过非法URL：{url}")
            continue

        # 网页抓取
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')

        # 文本提取
        text = soup.get_text()

        # 关键词句子识别
        sentences = sent_tokenize(text)
        keyword = "emotions"
        result_sentences = [sentence for sentence in sentences if keyword in sentence]

        # 获取当前时间
        now = datetime.datetime.now()

        # 创建文件夹名
        folder_name = f"{now.strftime('%Y%m%d_%H%M%S')}_{url.split('/')[-1].replace('?', '_')}"

        # 创建文件夹
        folder_path = os.path.join(save_path, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 保存文件
        wb = Workbook()
        ws = wb.active
        ws.title = "key sentence"
        ws['A1'] = "sentence"
        for i, sentence in enumerate(result_sentences):
            ws.cell(row=i+2, column=1).value = sentence
        wb.save(os.path.join(folder_path, f"key sentence_{folder_name}.xlsx"))

        print(f"已保存文件夹：{folder_name}")
    except Exception as e:
        print(f"错误：{e}")