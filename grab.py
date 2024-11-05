import requests
from bs4 import BeautifulSoup
import nltk
from nltk.tokenize import sent_tokenize
from openpyxl import Workbook
import datetime
import os


# 网页抓取
url = "https://dementiadiaries.org/entry/6104/a-moving-account-from-carol-of-crying-and-uncontrollable-emotions/?highlight=emotions"
response = requests.get(url)
soup = BeautifulSoup(response.content, 'html.parser')


# 文本提取
text = soup.get_text()


# 关键词句子识别
nltk.download('punkt')
sentences = sent_tokenize(text)
keyword = "emotions"
result_sentences = [sentence for sentence in sentences if keyword in sentence]


# 获取当前时间
now = datetime.datetime.now()

# 创建文件名
filename = f"key sentence_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"


# 指定文件保存路径
save_path = os.path.join(os.getcwd(), "data_outcome")


# 如果保存路径不存在，则创建
if not os.path.exists(save_path):
    os.makedirs(save_path)


# 保存文件
wb = Workbook()
ws = wb.active
ws.title = "key sentence"
ws['A1'] = "sentence"
for i, sentence in enumerate(result_sentences):
    ws.cell(row=i+2, column=1).value = sentence
wb.save(os.path.join(save_path, filename))